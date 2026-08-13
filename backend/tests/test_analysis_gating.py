"""An analysis-incomplete property must not expose default-workbook scores, returns, or
conclusions through the memo, tabs, or exports; a property with the required inputs still
receives normal analysis. Presentation/output gating only — the workbook is untouched."""
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook


def _import(client: TestClient, address: str = "12 Gating Way, Hudson, NY 12534") -> dict:
    return client.post("/api/properties/import", json={"raw_address": address}).json()


def _complete(client: TestClient) -> dict:
    """Import then supply the critical input so analysis becomes property-specific."""
    prop = _import(client, "34 Complete Rd, Hudson, NY 12534")
    client.put(f"/api/properties/{prop['id']}", json={"asking_price": 640000, "annual_taxes": 12000})
    client.post(f"/api/properties/{prop['id']}/underwrite")
    return client.get(f"/api/properties/{prop['id']}").json()


def _csv_row(client: TestClient, property_id: int) -> dict:
    text = client.get(f"/api/properties/{property_id}/exports/csv").content.decode()
    header, values = (line.split(",") for line in text.splitlines()[:2])
    return dict(zip(header, values))


def _xlsx(client: TestClient, property_id: int):
    return load_workbook(BytesIO(client.get(f"/api/properties/{property_id}/exports/xlsx").content))


# ---- Memo ----

def test_incomplete_memo_withholds_all_default_conclusions(client: TestClient) -> None:
    prop = _import(client)
    assert prop["analysis_incomplete"] is True
    memo = client.get(f"/api/properties/{prop['id']}/report").json()
    assert memo["analysis_incomplete"] is True
    # No overall/Bistate score, no DSCR/CoC/financial-strength conclusions.
    assert "score of" not in memo["executive_summary"].lower()
    assert memo["strengths"] == []
    assert memo["weaknesses"] == []
    blob = " ".join(memo["strengths"] + memo["weaknesses"] + [memo["executive_summary"]]).lower()
    for leak in ("debt service coverage", "1.25x", "cash-on-cash", "acquisition score", "benchmark", "70.5"):
        assert leak not in blob, leak
    # Default-workbook payloads are emptied.
    assert memo["financial_summary"] == {} and memo["projected_returns"] == {} and memo["assumptions_used"] == {}
    assert memo["cash_required"] == 0
    # Required inputs are named; real verified facts are preserved.
    assert "asking price" in " ".join(memo["required_inputs"]).lower()
    assert any("hudson" in fact.lower() for fact in memo["verified_facts"])


def test_complete_memo_receives_normal_analysis(client: TestClient) -> None:
    prop = _complete(client)
    assert prop["analysis_incomplete"] is False
    memo = client.get(f"/api/properties/{prop['id']}/report").json()
    assert memo["analysis_incomplete"] is False
    assert "score of" in memo["executive_summary"].lower()
    assert memo["strengths"]                       # real conclusions present
    assert memo["financial_summary"].get("total_cash_required")
    assert memo["cash_required"] > 0


# ---- Exports (CSV / XLSX / PDF) ----

def test_incomplete_csv_omits_default_score_and_irr(client: TestClient) -> None:
    prop = _import(client)
    row = _csv_row(client, prop["id"])
    assert row["analysis_state"] == "analysis incomplete"
    assert row["overall_score"] == "" and row["irr"] == ""


def test_complete_csv_includes_score_and_irr(client: TestClient) -> None:
    prop = _complete(client)
    row = _csv_row(client, prop["id"])
    assert row["analysis_state"] == "analysis complete"
    assert float(row["overall_score"]) > 0 and float(row["irr"]) != 0


def test_incomplete_xlsx_omits_workbook_output(client: TestClient) -> None:
    prop = _import(client)
    wb = _xlsx(client, prop["id"])
    assert "Workbook Output" not in wb.sheetnames
    assert "Analysis Incomplete" in wb.sheetnames
    summary = {r[0]: r[1] for r in wb["Property Summary"].iter_rows(values_only=True)}
    assert summary["Overall score"] == "Analysis incomplete"


def test_complete_xlsx_includes_workbook_output(client: TestClient) -> None:
    prop = _complete(client)
    wb = _xlsx(client, prop["id"])
    assert "Workbook Output" in wb.sheetnames
    metrics = {r[0]: r[1] for r in wb["Workbook Output"].iter_rows(values_only=True)}
    assert metrics.get("total_cash_required")


def test_incomplete_pdf_export_renders_and_uses_gated_memo(client: TestClient) -> None:
    from app.models.property import Property
    from app.services.acquisition import build_investment_memo
    prop = _import(client)
    assert client.get(f"/api/properties/{prop['id']}/exports/pdf").status_code == 200
    # The PDF is built from the memo, which is gated at the source.
    from app.database import SessionLocal
    with SessionLocal() as db:
        record = db.get(Property, prop["id"])
        assert build_investment_memo(record)["analysis_incomplete"] is True
        assert build_investment_memo(record)["strengths"] == []
